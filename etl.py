import argparse
import json
import os

import openpyxl
import pandas as pd

from config import (
    Thresholds,
    ORDER_HEADER_KEYWORDS,
    SUPPLIER_HEADER_KEYWORDS,
    ORDER_CANONICAL_MAP,
    SUPPLIER_CANONICAL_MAP,
)
from io_excel import sheet_is_orders, extract_table_with_provenance, canonicalize_columns
from suppliers import build_suppliers_clean, deduplicate_suppliers
from suppliers_master import build_suppliers_master
from certifications import split_certifications
from orders import clean_and_match_orders
from report import write_outputs, build_orders_supplier_cert_report
from odoo_export import write_odoo_outputs
from utils import now_iso


def _write_non_cdc_pack(
    *,
    out_dir: str,
    xlsx_path: str,
    df_sup_dedup: pd.DataFrame,
    df_certs: pd.DataFrame,
    df_orders_clean: pd.DataFrame,
    df_orders_supplier_cert_report: pd.DataFrame,
    norm_logs: list,
    match_warnings: list,
    unmatched_orders: list,
) -> None:
    """
    Produce the same reports but only for orders where JOB/CDC does NOT start with 'CDC'.
    Writes to: <out_dir>/non_cdc/
    """
    non_cdc_dir = os.path.join(out_dir, "non_cdc")
    os.makedirs(non_cdc_dir, exist_ok=True)

    if "job_cdc_is_cdc" not in df_orders_clean.columns:
        raise RuntimeError(
            "Missing column 'job_cdc_is_cdc' in orders_clean. "
            "Update orders.clean_and_match_orders() to add CDC classification columns."
        )

    # Filter orders (NON-CDC)
    df_orders_non_cdc = df_orders_clean[
        ~df_orders_clean["job_cdc_is_cdc"].fillna(False).astype(bool)
    ].copy()
    non_cdc_order_ids = set(df_orders_non_cdc["order_id"].astype(str))
    df_orders_supplier_cert_report_non_cdc = df_orders_supplier_cert_report[
        df_orders_supplier_cert_report["order_id"].astype(str).isin(non_cdc_order_ids)
    ].copy()

    # Filter match_warnings for those orders
    df_match_warnings_all = pd.DataFrame(match_warnings)
    if not df_match_warnings_all.empty and "order_id" in df_match_warnings_all.columns:
        df_match_warnings_non_cdc = df_match_warnings_all[
            df_match_warnings_all["order_id"].astype(str).isin(non_cdc_order_ids)
        ].copy()
    else:
        df_match_warnings_non_cdc = pd.DataFrame()

    # Filter unmatched_orders for those orders
    df_unmatched_orders_all = pd.DataFrame(unmatched_orders)
    if not df_unmatched_orders_all.empty and "order_id" in df_unmatched_orders_all.columns:
        df_unmatched_orders_non_cdc = df_unmatched_orders_all[
            df_unmatched_orders_all["order_id"].astype(str).isin(non_cdc_order_ids)
        ].copy()
    else:
        df_unmatched_orders_non_cdc = pd.DataFrame()

    # Filter normalization log to transformations for those orders
    df_norm_all = pd.DataFrame(norm_logs)
    if not df_norm_all.empty and {"entity", "record_id"}.issubset(df_norm_all.columns):
        df_norm_non_cdc = df_norm_all[
            (df_norm_all["entity"] == "order")
            & (df_norm_all["record_id"].astype(str).isin(non_cdc_order_ids))
        ].copy()
    else:
        df_norm_non_cdc = pd.DataFrame()

    # Write NON-CDC outputs (same filenames, separate folder)
    df_orders_non_cdc.to_csv(os.path.join(non_cdc_dir, "orders_clean.csv"), index=False)
    df_orders_supplier_cert_report_non_cdc.to_csv(
        os.path.join(non_cdc_dir, "orders_supplier_cert_report.csv"), index=False
    )
    df_match_warnings_non_cdc.to_csv(os.path.join(non_cdc_dir, "match_warnings.csv"), index=False)
    df_unmatched_orders_non_cdc.to_csv(os.path.join(non_cdc_dir, "unmatched_orders.csv"), index=False)
    df_norm_non_cdc.to_csv(os.path.join(non_cdc_dir, "normalization_log.csv"), index=False)

    # Optional: include suppliers/certs so the folder is a complete “pack”
    df_sup_dedup.to_csv(os.path.join(non_cdc_dir, "suppliers_clean.csv"), index=False)
    df_certs.to_csv(os.path.join(non_cdc_dir, "certifications_clean.csv"), index=False)

    summary_non_cdc = {
        "source_file": os.path.basename(xlsx_path),
        "scope": "orders where JOB/CDC does NOT start with CDC",
        "run_finished_at": now_iso(),
        "counts": {
            "orders_rows_clean": int(len(df_orders_non_cdc)),
            "orders_supplier_cert_report_rows": int(len(df_orders_supplier_cert_report_non_cdc)),
            "match_warnings": int(len(df_match_warnings_non_cdc)),
            "unmatched_orders": int(len(df_unmatched_orders_non_cdc)),
            "normalization_ops_orders": int(len(df_norm_non_cdc)),
        },
    }
    with open(os.path.join(non_cdc_dir, "summary_report.json"), "w", encoding="utf-8") as f:
        json.dump(summary_non_cdc, f, ensure_ascii=False, indent=2)



def _check_certifications_referential_integrity(
    *,
    df_sup_dedup: pd.DataFrame,
    df_certs: pd.DataFrame,
    out_dir: str,
) -> set[str]:
    supplier_ids = set(df_sup_dedup.get("supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip())
    cert_supplier_ids = set(df_certs.get("supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip())

    supplier_ids.discard("")
    cert_supplier_ids.discard("")

    orphan_ids = cert_supplier_ids - supplier_ids
    if orphan_ids:
        orphan_df = df_certs[df_certs["supplier_id"].astype(str).isin(orphan_ids)].copy()
        orphan_df.to_csv(os.path.join(out_dir, "orphan_certifications.csv"), index=False)
        raise RuntimeError(
            "Referential integrity error: certifications_clean contains supplier_id values "
            "not present in suppliers_clean. See output/orphan_certifications.csv for details."
        )

    orphan_path = os.path.join(out_dir, "orphan_certifications.csv")
    if os.path.exists(orphan_path):
        os.remove(orphan_path)

    return orphan_ids


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input_xlsx")
    ap.add_argument("--out", default="./output")
    ap.add_argument("--low", type=int, default=80)
    ap.add_argument("--high", type=int, default=92)

    # NEW: enable/disable the split pack from CLI
    ap.add_argument(
        "--split-non-cdc",
        action="store_true",
        help="Also produce a second report pack under <out>/non_cdc with only orders whose JOB/CDC does NOT start with 'CDC'.",
    )

    args = ap.parse_args()

    xlsx_path = os.path.abspath(args.input_xlsx)
    out_dir = os.path.abspath(args.out)
    thresholds = Thresholds(low=args.low, high=args.high)

    norm_logs = []
    match_warnings = []
    unmatched_orders = []
    unmatched_suppliers_parts = []
    suppliers_parts = []
    orders_parts = []
    supplier_sheet_diagnostics = []

    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sh in wb.sheetnames:
        df_preview = pd.read_excel(
            xlsx_path, sheet_name=sh, header=None, engine="openpyxl", nrows=40
        )

        if sheet_is_orders(df_preview, ORDER_HEADER_KEYWORDS):
            df_raw, _ = extract_table_with_provenance(
                xlsx_path, sh, "orders", ORDER_HEADER_KEYWORDS, SUPPLIER_HEADER_KEYWORDS
            )
            if df_raw.empty:
                continue
            df_can = canonicalize_columns(df_raw, ORDER_CANONICAL_MAP)
            orders_parts.append(df_can)
        else:
            df_raw, header_idx = extract_table_with_provenance(
                xlsx_path, sh, "suppliers", ORDER_HEADER_KEYWORDS, SUPPLIER_HEADER_KEYWORDS
            )
            if df_raw.empty:
                continue
            header_norm = [str(c) for c in df_raw.columns if str(c) not in {"source_file", "source_sheet", "source_row"}]
            mapped = []
            for c in header_norm:
                tgt = SUPPLIER_CANONICAL_MAP.get(c)
                if tgt:
                    mapped.append(f"{c}->{tgt}")
            supplier_sheet_diagnostics.append({
                "source_file": os.path.basename(xlsx_path),
                "source_sheet": sh,
                "header_row_index_0_based": header_idx,
                "header_row_excel_1_based": header_idx + 1,
                "raw_headers_normalized": json.dumps(header_norm, ensure_ascii=False),
                "mapped_columns": json.dumps(mapped, ensure_ascii=False),
                "expiry_field_mapped": any(m.endswith("->cert_expiry_raw") for m in mapped),
            })
            df_can = canonicalize_columns(df_raw, SUPPLIER_CANONICAL_MAP)
            df_clean, df_bad = build_suppliers_clean(
                df_can, supplier_category=sh, norm_logs=norm_logs
            )
            suppliers_parts.append(df_clean)
            if not df_bad.empty:
                unmatched_suppliers_parts.append(df_bad)

    df_sup_all = (
        pd.concat(suppliers_parts, ignore_index=True)
        if suppliers_parts
        else pd.DataFrame()
    )
    df_orders_raw = (
        pd.concat(orders_parts, ignore_index=True) if orders_parts else pd.DataFrame()
    )
    df_unmatched_sup = (
        pd.concat(unmatched_suppliers_parts, ignore_index=True)
        if unmatched_suppliers_parts
        else pd.DataFrame()
    )

    df_sup_dedup = deduplicate_suppliers(
        df_sup_all, match_warnings, low_threshold=thresholds.low
    )
    df_certs, cert_warnings, cert_tokens_dropped_nan_like = split_certifications(df_sup_dedup)
    orphan_ids = _check_certifications_referential_integrity(
        df_sup_dedup=df_sup_dedup,
        df_certs=df_certs,
        out_dir=out_dir,
    )
    df_orders_clean = clean_and_match_orders(
        df_orders_raw,
        df_sup_dedup,
        thresholds.low,
        thresholds.high,
        norm_logs,
        match_warnings,
        unmatched_orders,
    )
    df_orders_supplier_cert_report = build_orders_supplier_cert_report(
        orders=df_orders_clean,
        suppliers=df_sup_dedup,
        certs=df_certs,
    )

    orders_cdc_count = int(
        df_orders_clean.get("job_cdc_is_cdc", pd.Series(dtype=bool)).fillna(False).astype(bool).sum()
    )
    orders_total_count = int(len(df_orders_clean))
    orders_non_cdc_count = int(max(0, orders_total_count - orders_cdc_count))

    (
        df_suppliers_master,
        df_suppliers_master_unmatched,
        df_suppliers_registry_not_used,
        df_registry_to_master_mapping,
    ) = build_suppliers_master(
        df_orders_clean=df_orders_clean,
        df_registry_suppliers=df_sup_dedup,
        df_certs=df_certs,
        low=thresholds.low,
        high=thresholds.high,
        match_warnings=match_warnings,
    )

    supplier_uid_by_key = (
        df_suppliers_master[["supplier_name_key", "supplier_external_uid"]]
        .drop_duplicates(subset=["supplier_name_key"], keep="first")
    ) if not df_suppliers_master.empty else pd.DataFrame(columns=["supplier_name_key", "supplier_external_uid"])
    df_orders_clean = df_orders_clean.merge(
        supplier_uid_by_key,
        how="left",
        left_on="order_supplier_key",
        right_on="supplier_name_key",
    ).drop(columns=["supplier_name_key"], errors="ignore")
    df_orders_clean["supplier_external_uid"] = df_orders_clean["supplier_external_uid"].fillna("")

    summary = {
        "source_file": os.path.basename(xlsx_path),
        "run_started_at": now_iso(),
        "run_finished_at": now_iso(),
        "thresholds": {"low": thresholds.low, "high": thresholds.high},
        "counts": {
            "sheets_total": len(wb.sheetnames),
            "orders_rows_read": int(len(df_orders_raw)),
            "orders_rows_clean": int(len(df_orders_clean)),
            "orders_rows_clean_total": orders_total_count,
            "orders_rows_clean_cdc": orders_cdc_count,
            "orders_rows_clean_non_cdc": orders_non_cdc_count,
            "orders_supplier_cert_report_rows": int(len(df_orders_supplier_cert_report)),
            "suppliers_rows_read": int(len(df_sup_all)),
            "suppliers_rows_clean": int(len(df_sup_dedup)),
            "suppliers_dedup_merged": int(max(0, len(df_sup_all) - len(df_sup_dedup))),
            "certifications_rows": int(len(df_certs)),
            "suppliers_rows_with_expiry_raw_nonempty": int((df_sup_dedup.get("cert_expiry_raw", pd.Series(dtype=str)).fillna("").astype(str).str.strip() != "").sum()),
            "cert_rows_with_expiry_parsed": int((df_certs.get("expiry_date", pd.Series(dtype=str)).fillna("").astype(str).str.strip() != "").sum()),
            "cert_rows_with_expiry_unparseable": int(sum(1 for w in cert_warnings if w.get("warning_type") == "cert_expiry_unparseable")),
            "cert_rows_with_expiry_missing": int(sum(1 for w in cert_warnings if w.get("warning_type") == "cert_expiry_missing")),
            "cert_tokens_dropped_nan_like": int(cert_tokens_dropped_nan_like),
            "supplier_ids_in_suppliers": int(df_sup_dedup.get("supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()),
            "supplier_ids_in_certs": int(df_certs.get("supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()),
            "orphan_cert_supplier_ids": int(len(orphan_ids)),
            "normalization_ops": int(len(norm_logs)),
            "match_warnings": int(len(match_warnings)),
            "unmatched_orders": int(len(unmatched_orders)),
            "unmatched_suppliers": int(len(df_unmatched_sup)) if not df_unmatched_sup.empty else 0,
            "suppliers_master_count": int(len(df_suppliers_master)),
            "suppliers_master_matched_registry_count": int((df_suppliers_master.get("matched_registry_supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip() != "").sum()),
            "suppliers_master_unmatched_registry_count": int((df_suppliers_master.get("matched_registry_supplier_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip() == "").sum()),
            "suppliers_master_with_certs_count": int((df_suppliers_master.get("has_certifications", pd.Series(dtype=bool)).fillna(False) == True).sum()),
            "suppliers_master_without_certs_count": int((df_suppliers_master.get("has_certifications", pd.Series(dtype=bool)).fillna(False) == False).sum()),
        },
    }

    missing_master_for_order_keys = set(df_orders_clean["order_supplier_key"].fillna("").astype(str)) - set(df_suppliers_master["supplier_name_key"].fillna("").astype(str))
    if missing_master_for_order_keys:
        raise RuntimeError("Data integrity error: some order_supplier_key values are missing in suppliers_master")

    if (df_orders_clean["supplier_external_uid"].astype(str).str.strip() == "").any():
        raise RuntimeError("Data integrity error: at least one order does not reference supplier_master.external_uid")

    # Main (full) outputs
    write_outputs(
        out_dir=out_dir,
        suppliers=df_sup_dedup,
        certs=df_certs,
        orders=df_orders_clean,
        orders_supplier_cert_report=df_orders_supplier_cert_report,
        norm_logs=norm_logs,
        match_warnings=match_warnings,
        unmatched_orders=unmatched_orders,
        unmatched_suppliers=df_unmatched_sup,
        cert_warnings=cert_warnings,
        summary=summary,
        suppliers_master=df_suppliers_master,
        suppliers_registry_not_used_in_orders=df_suppliers_registry_not_used,
        suppliers_master_unmatched_registry=df_suppliers_master_unmatched,
        registry_to_master_mapping=df_registry_to_master_mapping,
    )

    odoo_counts = write_odoo_outputs(
        out_dir=out_dir,
        suppliers=df_suppliers_master,
        certs=df_certs,
        orders=df_orders_clean,
        registry_to_master_mapping=df_registry_to_master_mapping,
        run_ts=now_iso(),
    )
    summary["counts"].update(odoo_counts)
    summary["counts"]["odoo_fornitori_rows_full"] = int(odoo_counts.get("odoo_fornitori_rows", 0))
    summary["counts"]["odoo_certificazioni_rows_full"] = int(odoo_counts.get("odoo_certificazioni_rows", 0))
    summary["counts"]["odoo_ordini_rows_full"] = int(odoo_counts.get("odoo_ordini_rows", 0))

    if args.split_non_cdc:
        if "job_cdc_is_cdc" not in df_orders_clean.columns:
            raise RuntimeError(
                "Missing column 'job_cdc_is_cdc' in orders_clean. "
                "Update orders.clean_and_match_orders() to add CDC classification columns."
            )

        df_orders_non_cdc = df_orders_clean[
            ~df_orders_clean["job_cdc_is_cdc"].fillna(False).astype(bool)
        ].copy()
        non_cdc_supplier_uids = set(
            df_orders_non_cdc.get("supplier_external_uid", pd.Series(dtype=str))
            .fillna("")
            .astype(str)
            .str.strip()
        )
        non_cdc_supplier_uids.discard("")
        df_suppliers_master_non_cdc = df_suppliers_master[
            df_suppliers_master.get("supplier_external_uid", pd.Series(dtype=str))
            .fillna("")
            .astype(str)
            .isin(non_cdc_supplier_uids)
        ].copy()

        df_registry_to_master_mapping_non_cdc = df_registry_to_master_mapping[
            df_registry_to_master_mapping.get("supplier_external_uid", pd.Series(dtype=str))
            .fillna("")
            .astype(str)
            .isin(non_cdc_supplier_uids)
        ].copy()

        odoo_non_cdc_counts = write_odoo_outputs(
            out_dir=os.path.join(out_dir, "non_cdc"),
            suppliers=df_suppliers_master_non_cdc,
            certs=df_certs,
            orders=df_orders_non_cdc,
            registry_to_master_mapping=df_registry_to_master_mapping_non_cdc,
            run_ts=now_iso(),
        )
        summary["counts"].update(
            {f"odoo_non_cdc_{k.removeprefix('odoo_')}": v for k, v in odoo_non_cdc_counts.items()}
        )
        summary["counts"]["odoo_fornitori_rows_non_cdc"] = int(odoo_non_cdc_counts.get("odoo_fornitori_rows", 0))
        summary["counts"]["odoo_certificazioni_rows_non_cdc"] = int(odoo_non_cdc_counts.get("odoo_certificazioni_rows", 0))
        summary["counts"]["odoo_ordini_rows_non_cdc"] = int(odoo_non_cdc_counts.get("odoo_ordini_rows", 0))

    with open(os.path.join(out_dir, "summary_report.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    pd.DataFrame(supplier_sheet_diagnostics).to_csv(
        os.path.join(out_dir, "supplier_header_diagnostics.csv"), index=False
    )

    # Optional split pack: NON-CDC
    if args.split_non_cdc:
        _write_non_cdc_pack(
            out_dir=out_dir,
            xlsx_path=xlsx_path,
            df_sup_dedup=df_sup_dedup,
            df_certs=df_certs,
            df_orders_clean=df_orders_clean,
            df_orders_supplier_cert_report=df_orders_supplier_cert_report,
            norm_logs=norm_logs,
            match_warnings=match_warnings,
            unmatched_orders=unmatched_orders,
        )

    print(f"Done. Outputs in: {out_dir}")
    print(
        "Orders split counters: "
        f"total={orders_total_count}, "
        f"cdc={orders_cdc_count}, "
        f"non_cdc={orders_non_cdc_count}"
    )
    if args.split_non_cdc:
        print(f"NON-CDC pack in: {os.path.join(out_dir, 'non_cdc')}")


if __name__ == "__main__":
    main()
